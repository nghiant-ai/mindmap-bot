"""
Excel Mind Map Generator
Generates mindmap in Excel format for easy editing by non-technical users
"""
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from app.core.logging import get_logger
from app.utils.date_utils import get_vietnam_time
from app.services.mindmap.models import MindMapNode, MindMapTheme

logger = get_logger(__name__)


class ExcelGenerator:
    """
    Generate mind map in Excel format

    Benefits:
    - Easy to edit for non-technical users
    - Visual hierarchy with indentation
    - Can add/remove/reorder nodes easily
    - Color coding by level
    - Can import back to EdrawMind from Excel

    Format:
    | Level | Node Name | Note | Path |
    |-------|-----------|------|------|
    | 0     | Root      | ...  | Root |
    | 1     | Branch 1  | ...  | Root > Branch 1 |
    | 2     | Sub 1.1   | ...  | Root > Branch 1 > Sub 1.1 |
    """

    # Color scheme for different levels
    LEVEL_COLORS = {
        0: "4472C4",  # Blue - Root
        1: "70AD47",  # Green - Main branches
        2: "FFC000",  # Orange - Sub-branches
        3: "5B9BD5",  # Light blue
        4: "A5A5A5",  # Gray
    }

    def __init__(self):
        self.row_num = 1  # Track current row

    async def generate(
        self,
        root: MindMapNode,
        title: str,
        theme: MindMapTheme
    ) -> BytesIO:
        """
        Generate Excel mind map

        Args:
            root: Root node of the mind map
            title: Mind map title
            theme: Color theme (not used but kept for interface consistency)

        Returns:
            BytesIO buffer with Excel content
        """
        # Create workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Mind Map"

        # Add metadata sheet
        self._add_metadata_sheet(wb, title)

        # Setup main sheet
        ws = wb["Mind Map"]
        self._setup_headers(ws)

        # Reset row counter
        self.row_num = 2  # Start after header

        # Build Excel structure
        self._build_excel_tree(ws, root, level=0, path=root.name)

        # Auto-adjust column widths
        self._adjust_column_widths(ws)

        # Freeze header row
        ws.freeze_panes = "A2"

        # Save to BytesIO
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        logger.info(f"Excel generated: {title} ({self.row_num - 1} nodes)")
        return buffer

    def _setup_headers(self, ws):
        """Setup header row with styling"""
        headers = ["Level", "Node Name", "Note", "Path"]

        # Header style
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=12)
        border = Border(
            bottom=Side(style='medium', color='000000')
        )

        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border

    def _build_excel_tree(self, ws, node: MindMapNode, level: int, path: str):
        """
        Recursively build Excel tree structure

        Args:
            ws: Worksheet
            node: Current MindMapNode
            level: Current depth level
            path: Breadcrumb path (e.g., "Root > Branch 1 > Sub 1.1")
        """
        # Get color for this level
        color = self.LEVEL_COLORS.get(level, "FFFFFF")

        # Level column
        level_cell = ws.cell(row=self.row_num, column=1)
        level_cell.value = level
        level_cell.alignment = Alignment(horizontal='center')
        level_cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")

        # Node name column (with indentation based on level)
        name_cell = ws.cell(row=self.row_num, column=2)
        indent = "  " * level  # Two spaces per level
        name_cell.value = f"{indent}{node.name}"
        name_cell.font = Font(bold=(level == 0), size=11 if level == 0 else 10)

        # Note column
        note_cell = ws.cell(row=self.row_num, column=3)
        note_cell.value = node.note if node.note else ""
        note_cell.alignment = Alignment(wrap_text=True)

        # Path column (breadcrumb)
        path_cell = ws.cell(row=self.row_num, column=4)
        path_cell.value = path
        path_cell.font = Font(size=9, color="666666")

        self.row_num += 1

        # Recursively add children
        for child in node.children:
            child_path = f"{path} > {child.name}"
            self._build_excel_tree(ws, child, level + 1, child_path)

    def _adjust_column_widths(self, ws):
        """Auto-adjust column widths for better readability"""
        column_widths = {
            1: 8,   # Level
            2: 40,  # Node Name
            3: 50,  # Note
            4: 60,  # Path
        }

        for col_num, width in column_widths.items():
            ws.column_dimensions[get_column_letter(col_num)].width = width

    def _add_metadata_sheet(self, wb, title: str):
        """Add metadata sheet with info about the mindmap"""
        meta_ws = wb.create_sheet("Info", 0)

        metadata = [
            ["Mind Map Title", title],
            ["Generated By", "Mindmap Bot"],
            ["Generated At", get_vietnam_time().strftime('%Y-%m-%d %H:%M:%S')],
            ["Format", "Excel (.xlsx)"],
            ["", ""],
            ["How to use:", ""],
            ["1. Edit 'Node Name' and 'Note' columns directly", ""],
            ["2. Add new rows to add nodes (maintain Level structure)", ""],
            ["3. Delete rows to remove nodes", ""],
            ["4. Keep Level column consistent (0=Root, 1=Branch, 2=Sub-branch, etc.)", ""],
            ["5. Save and use for reference or import back to EdrawMind", ""],
        ]

        for row_num, row_data in enumerate(metadata, 1):
            meta_ws.cell(row=row_num, column=1).value = row_data[0]
            if len(row_data) > 1 and row_data[1]:
                meta_ws.cell(row=row_num, column=2).value = row_data[1]

        # Style metadata
        meta_ws.column_dimensions['A'].width = 20
        meta_ws.column_dimensions['B'].width = 60

        # Bold the labels
        for row in range(1, 5):
            meta_ws.cell(row=row, column=1).font = Font(bold=True)


# Global instance
excel_generator = ExcelGenerator()
